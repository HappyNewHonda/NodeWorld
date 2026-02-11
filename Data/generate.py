"""
Excel(.xlsm) → JSON + C# クラス 変換ツール
元のVBAマクロ「JSON出力ツール」のPython版

使い方:
  python generate.py
"""

import json
import os
import re
import warnings
from pathlib import Path
from dataclasses import dataclass, field
from typing import Optional

import openpyxl
from jinja2 import Template

warnings.filterwarnings("ignore", category=UserWarning)

# ====== 設定 ======
EXCEL_FILE = "データ.xlsm"
TOOL_SHEET = "JSON出力ツール"
CONVERSION_STR = "-"
CONVERSION_TEXT = "-1"
START_MARKER = "ここから"
END_MARKER = "ここまで"
ADD_DATA_MARKER = "追加読み込み"
KEY_DATA_MARKER = "キーデータ"
COMMENT_PREFIX = "//"
# ==================


# ============================================================
#  色の取得・比較
# ============================================================
def get_cell_color(cell):
    """セルの背景色を比較可能な形式で返す"""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    fg = fill.fgColor
    if fg is None or fg.type is None:
        return None

    if fg.type == "rgb" and fg.rgb and fg.rgb != "00000000":
        return ("rgb", str(fg.rgb))
    elif fg.type == "theme":
        theme = fg.theme
        tint = round(fg.tint, 4) if fg.tint else 0.0
        return ("theme", theme, tint)
    elif fg.type == "indexed":
        return ("indexed", fg.indexed)
    return None


def colors_match(c1, c2):
    """2つの色が同じかどうかを判定"""
    if c1 is None or c2 is None:
        return False
    return c1 == c2


# ============================================================
#  フィールド定義
# ============================================================
@dataclass
class FieldDef:
    name: str
    type_name: str
    caption: str = ""
    is_array: bool = False
    column_indices: list = field(default_factory=list)  # 1始まり


# ============================================================
#  シートからデータを読み取る
# ============================================================
def read_sheet_data(wb, sheet_name, target_color, target_columns=None):
    """
    シートからフィールド定義とデータを読み取る

    Returns:
        fields: list[FieldDef]
        rows: list[dict]  各行のデータ
        add_data: list[str]  追加読み込みデータ
        key_data: list[str]  キーデータ
        is_primary_key: bool
    """
    # シートを探す
    actual_name = None
    for name in wb.sheetnames:
        if name == sheet_name or name.upper() == sheet_name.upper():
            actual_name = name
            break
    if not actual_name:
        print(f"  警告: シート「{sheet_name}」が見つかりません。スキップします。")
        return None, None, None, None, False

    ws = wb[actual_name]

    # 「ここから」「ここまで」を探す
    start_row = None
    end_row = None
    for r in range(1, ws.max_row + 1):
        val = ws.cell(row=r, column=1).value
        if val == START_MARKER and start_row is None:
            start_row = r
        elif val == END_MARKER and start_row is not None:
            end_row = r
            break

    if start_row is None:
        print(f"  警告: シート「{actual_name}」に「{START_MARKER}」が見つかりません。")
        return None, None, None, None, False
    if end_row is None:
        print(f"  警告: シート「{actual_name}」に「{END_MARKER}」が見つかりません。")
        return None, None, None, None, False

    head_row = start_row + 1       # 説明行（キャプション）
    def_row = start_row + 2        # フィールド定義行
    data_start_row = start_row + 3  # データ開始行

    # データが空の場合
    if end_row - 1 < data_start_row:
        return [], [], [], [], False

    # ---- 対象列の特定 ----
    use_color = (target_columns is None)
    all_columns = []  # (col_index, caption, field_def_str) のリスト

    col = 1
    while True:
        caption = ws.cell(row=head_row, column=col).value
        if caption is None or (isinstance(caption, str) and caption.strip() == ""):
            break

        field_def_str = ws.cell(row=def_row, column=col).value

        if use_color:
            col_color = get_cell_color(ws.cell(row=head_row, column=col))
            if colors_match(col_color, target_color):
                all_columns.append((col, str(caption), field_def_str))
        else:
            # 列名指定の場合
            col_letter = openpyxl.utils.get_column_letter(col)
            if col_letter in target_columns:
                all_columns.append((col, str(caption), field_def_str))

        col += 1

    # ---- フィールド定義を構築（同名フィールドは配列化） ----
    fields = []
    field_map = {}  # name -> FieldDef

    for col_idx, caption, def_str in all_columns:
        if def_str is None or (isinstance(def_str, str) and def_str.strip() == ""):
            continue

        def_str = str(def_str).strip()
        parts = def_str.split(":")
        fname = parts[0]
        ftype = parts[-1] if len(parts) > 1 else "string"

        # 型名から[]を除去（配列は同名カラムで判定）
        is_explicit_array = ftype.endswith("[]")
        if is_explicit_array:
            ftype = ftype[:-2]

        if fname in field_map:
            # 既存フィールドに列を追加 → 配列
            field_map[fname].column_indices.append(col_idx)
            field_map[fname].is_array = True
        else:
            fd = FieldDef(
                name=fname,
                type_name=ftype,
                caption=caption,
                is_array=is_explicit_array,
                column_indices=[col_idx]
            )
            fields.append(fd)
            field_map[fname] = fd

    # ---- データ行を読み取る ----
    rows = []
    for r in range(data_start_row, end_row):
        first_val = ws.cell(row=r, column=1).value
        if first_val is not None and str(first_val).startswith(COMMENT_PREFIX):
            continue

        row_data = {}
        for fd in fields:
            is_multi_column = len(fd.column_indices) > 1  # 実際に複数列あるか

            if is_multi_column:
                # 複数列 → 各列の値を集めて配列
                values = []
                for ci in fd.column_indices:
                    v = ws.cell(row=r, column=ci).value
                    v = convert_value(v, fd.type_name)
                    values.append(v)
                row_data[fd.name] = values
            else:
                # 単一列（is_arrayがTrueでも列は1つ = セルの中身が既に配列文字列）
                v = ws.cell(row=r, column=fd.column_indices[0]).value

                if v is None:
                    if fd.type_name in ("int", "bool"):
                        row_data[fd.name] = 0
                    elif fd.type_name == "float":
                        row_data[fd.name] = 0.0
                    else:
                        row_data[fd.name] = ""
                else:
                    s = str(v)
                    if s == CONVERSION_STR:
                        s = CONVERSION_TEXT

                    if fd.type_name == "string":
                        row_data[fd.name] = s
                    else:
                        # string以外: 数値変換を試みる、失敗したらそのまま（raw値）
                        try:
                            if fd.type_name in ("int", "bool"):
                                row_data[fd.name] = int(float(s))
                            elif fd.type_name == "float":
                                row_data[fd.name] = float(s)
                            else:
                                row_data[fd.name] = s
                        except (ValueError, TypeError):
                            # "[3]" や "[2,5]" などはそのまま文字列として保持
                            row_data[fd.name] = s

        rows.append(row_data)


    # ---- 追加読み込み ----
    add_data = []
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == ADD_DATA_MARKER:
            c = 1
            while True:
                v = ws.cell(row=r + 1, column=c).value
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    break
                # エスケープシーケンスの変換
                text = str(v)
                # \n の後の空白を \n\t に
                text = re.sub(r'\\n\s', r'\\n\\t', text)
                # \t の後の空白を \t\t に
                while re.search(r'\\t\s', text):
                    text = re.sub(r'\\t\s', r'\\t\\t', text)
                text = text.replace('\\n', '\n').replace('\\t', '\t')
                add_data.append(text)
                c += 1
            break

    # ---- キーデータ ----
    key_data = []
    is_primary_key = False
    for r in range(1, ws.max_row + 1):
        if ws.cell(row=r, column=1).value == KEY_DATA_MARKER:
            c = 1
            while True:
                v = ws.cell(row=r + 1, column=c).value
                if v is None or (isinstance(v, str) and v.strip() == ""):
                    break
                text = str(v).strip().replace(" ", "")

                # [PrimaryKey] / [Only] の判定
                if "[PrimaryKey]" in text or "[Only]" in text:
                    is_primary_key = True
                    text = text.replace("[PrimaryKey]", "").replace("[Only]", "")

                key_data.append(text)
                c += 1
            break

    return fields, rows, add_data, key_data, is_primary_key


def convert_value(value, type_name):
    """値を型に応じて変換する（VBAと同じ挙動）"""
    if value is None:
        if type_name == "int":
            return 0
        elif type_name == "bool":
            return 0
        elif type_name == "float":
            return 0.0
        return ""

    s = str(value)

    # コンバージョン文字の変換
    if s == CONVERSION_STR:
        s = CONVERSION_TEXT

    if type_name == "string":
        return s
    elif type_name == "int":
        # VBAでは IsNumeric チェックの上でそのまま出力
        # セルの値が "[3]" のような文字列ならそのまま返す（raw JSON埋め込み用）
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return 0
    elif type_name == "bool":
        try:
            return int(float(s))
        except (ValueError, TypeError):
            return 0
    elif type_name == "float":
        try:
            return float(s)
        except (ValueError, TypeError):
            return 0.0
    else:
        # string以外の未知の型（Colorなど）もそのまま
        return s


# ============================================================
#  JSON出力
# ============================================================
def generate_json(fields, rows):
    """JSON文字列を生成する（VBAと同じ方式で文字列を直接組み立てる）"""
    row_strings = []

    for row in rows:
        field_strings = []
        for fd in fields:
            val = row.get(fd.name)
            field_str = format_json_value(fd, val)
            field_strings.append(f'"{fd.name}":{field_str}')
        row_strings.append("{" + ",".join(field_strings) + "}")

    return '{"data":[' + ",".join(row_strings) + ']}'


def format_json_value(fd, val):
    """VBAと同じ方式でJSON値を出力する"""
    is_multi_column = len(fd.column_indices) > 1  # ← is_arrayではなく実際の列数で判定

    if is_multi_column:
        # 実際に複数列 → Python側で配列化
        if not isinstance(val, list):
            val = [val]
        parts = []
        for v in val:
            if fd.type_name == "string":
                parts.append(f'"{v}"')
            else:
                parts.append(str(v))
        return "[" + ",".join(parts) + "]"
    else:
        # 単一列 → そのまま出力
        if fd.type_name == "string":
            return f'"{val}"'
        else:
            # val が "[3]" や "[2,5]" ならそのまま、数値ならそのまま
            return str(val)


# ============================================================
#  C# コード生成
# ============================================================
def to_pascal_case(name):
    """snake_case → PascalCase"""
    parts = name.split("_")
    result = ""
    for part in parts:
        if part.isdigit():
            continue
        if part:
            result += part[0].upper() + part[1:]
    return result


CS_TEMPLATE = Template('''\
//--------------------------------------------------------------------------------------
// Generated by ExcelToPython From {{ book_name }}
//--------------------------------------------------------------------------------------
using System;
{% if has_keys -%}
using System.Collections.Generic;
using System.Linq;
using UnityEngine;
{% endif %}
namespace Data.Master
{
\t/// <summary>
\t/// {{ sheet_name }}データ
\t/// </summary>
\t[Serializable]
\tpublic partial class {{ class_name }}s{{ implements }}
\t{
\t\t/// <summary>
\t\t/// {{ class_name }} の配列データ。
\t\t/// </summary>
\t\tpublic {{ class_name }}[] data;
{% if has_keys %}
{% for ku in key_utils %}
\t\t/// <summary>
\t\t/// アクセスユーティリティ: {{ ku.name }}（キーで {{ class_name }}{{ "[]" if not is_primary_key else "" }} を引けます）。
\t\t/// </summary>
{% if is_primary_key %}
\t\tpublic Dictionary<{{ ku.key_type }}, {{ class_name }}> {{ ku.name }} { get; private set; }
{% else %}
\t\tpublic Dictionary<{{ ku.key_type }}, {{ class_name }}[]> {{ ku.name }} { get; private set; }
{% endif %}
{% endfor %}

\t\tpublic void OnBeforeSerialize()
\t\t{
\t\t}

\t\tpublic void OnAfterDeserialize()
\t\t{
{% for ku in key_utils %}
{{ ku.deserialize_code }}
{% endfor %}
\t\t}
{% endif %}
\t}

\t/// <summary>
\t/// {{ sheet_name }}のデータモデル
\t/// </summary>
\t[Serializable]
\tpublic partial class {{ class_name }}
\t{
{% for f in fields %}
\t\t/// <summary>
\t\t/// {{ f.caption }}
\t\t/// </summary>
\t\tpublic {{ f.cs_type }} {{ f.name }};
{% endfor %}

\t\tpublic {{ class_name }} Clone()
\t\t{
\t\t\treturn ({{ class_name }}) MemberwiseClone();
\t\t}
\t}
{% for add_line in add_data %}
\t{{ add_line }}
{% endfor %}
}
''')


def generate_csharp(fields, rows, sheet_name, file_name, key_data, is_primary_key, add_data, book_name):
    """C#クラスファイルの文字列を生成する"""
    class_name = to_pascal_case(file_name)
    has_keys = len(key_data) > 0

    # フィールド情報
    cs_fields = []
    field_type_map = {}  # name -> type_name (キー型の解決用)
    for fd in fields:
        cs_type = fd.type_name
        if fd.is_array or len(fd.column_indices) > 1:
            cs_type = fd.type_name + "[]"

        cs_fields.append({
            "name": fd.name,
            "cs_type": cs_type,
            "caption": fd.caption if fd.caption else f"フィールド: {fd.name}（{cs_type}）"
        })
        field_type_map[fd.name] = fd.type_name

    # キーユーティリティ
    key_utils = []
    for kd in key_data:
        key_parts = [k.strip() for k in kd.split(",")]
        util_name = "Select" + "And".join(to_pascal_case(k) if "_" in k else (k[0].upper() + k[1:]) for k in key_parts)

        if len(key_parts) == 1:
            key_type = field_type_map.get(key_parts[0], "string")
        else:
            key_type = "string"

        # デシリアライズコード生成
        deserialize_code = generate_deserialize_code(
            util_name, key_parts, field_type_map, is_primary_key
        )

        key_utils.append({
            "name": util_name,
            "key_type": key_type,
            "deserialize_code": deserialize_code
        })

    implements = ""
    if has_keys:
        implements = " : ISerializationCallbackReceiver"

    result = CS_TEMPLATE.render(
        book_name=book_name,
        sheet_name=sheet_name,
        class_name=class_name,
        has_keys=has_keys,
        is_primary_key=is_primary_key,
        implements=implements,
        key_utils=key_utils,
        fields=cs_fields,
        add_data=add_data
    )

    # タブ文字の正規化（Jinja2が空白に変換する場合があるため）
    result = result.replace('\t', '\t')

    return result


def generate_deserialize_code(util_name, key_parts, field_type_map, is_primary_key):
    """OnAfterDeserialize内のコードを生成"""
    indent = "\t\t\t"

    if len(key_parts) == 1:
        key = key_parts[0]
        if is_primary_key:
            code = f"{indent}{util_name} = data.Select(x => x.{key}).Distinct().ToDictionary(key => key, val => data.First(x => x.{key} == val));"
        else:
            code = f"{indent}{util_name} = data.Select(x => x.{key}).Distinct().ToDictionary(key => key, val => data.Where(x => x.{key} == val).ToArray());"
    else:
        # 複合キー
        format_parts = []
        for i in range(len(key_parts)):
            format_parts.append("{" + str(i) + "}")
        format_str = "_".join(format_parts)

        select_args = ", ".join(f"x.{k}" for k in key_parts)
        select_expr = f'string.Format("{format_str}", {select_args})'

        where_conditions = []
        for i, k in enumerate(key_parts):
            ktype = field_type_map.get(k, "string")
            if ktype != "string":
                where_conditions.append(f"x.{k} == {ktype}.Parse(val.Split('_')[{i}])")
            else:
                where_conditions.append(f"x.{k} == val.Split('_')[{i}]")

        where_expr = " && ".join(where_conditions)

        if is_primary_key:
            code = f"{indent}{util_name} = data.Select(x => {select_expr}).Distinct().ToDictionary(key => key, val => data.First(x => {where_expr}));"
        else:
            code = f"{indent}{util_name} = data.Select(x => {select_expr}).Distinct().ToDictionary(key => key, val => data.Where(x => {where_expr}).ToArray());"

    return code


# ============================================================
#  ファイル保存
# ============================================================
def save_file(directory, filename, content, encoding="utf-8-sig"):
    """ファイルを保存する"""
    os.makedirs(directory, exist_ok=True)
    filepath = os.path.join(directory, filename)
    with open(filepath, "w", encoding=encoding) as f:
        f.write(content)
    return filepath


# ============================================================
#  メイン処理
# ============================================================
def main():
    print(f"ファイルを読み込み中: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    if TOOL_SHEET not in wb.sheetnames:
        print(f"エラー: シート「{TOOL_SHEET}」が見つかりません。")
        return

    ws = wb[TOOL_SHEET]
    base_dir = os.path.dirname(os.path.abspath(EXCEL_FILE))
    json_dir = os.path.join(base_dir, "json")

    # 設定を読み込み
    row = 2
    processed = 0
    skipped = 0

    while ws.cell(row=row, column=1).value is not None:
        a_val = str(ws.cell(row=row, column=1).value).strip()
        b_val = str(ws.cell(row=row, column=2).value).strip() if ws.cell(row=row, column=2).value else ""

        # C列に色がついていたらスキップ
        c_color = get_cell_color(ws.cell(row=row, column=3))
        if c_color is not None:
            print(f"  スキップ: {a_val}（C列に色あり）")
            row += 1
            skipped += 1
            continue

        # 設定の色を取得
        target_color = get_cell_color(ws.cell(row=row, column=1))

        # "+"で複数シート結合に対応
        sheet_specs = a_val.split("+")
        file_names = [s.strip() for s in b_val.replace("\r", "").replace("\n", "").replace(" ", "").split("+")]
        output_file_name = file_names[0]

        print(f"\n処理中: {a_val} → {output_file_name}")

        all_fields = None
        all_rows = []

        for spec in sheet_specs:
            spec = spec.strip().replace("\r", "").replace("\n", "").replace(" ", "")

            # "!"による列指定の判定
            target_columns = None
            if "!" in spec:
                parts = spec.split("!", 1)
                sheet_name = parts[0].strip()
                target_columns = [c.strip().upper() for c in parts[1].split(",")]
            else:
                sheet_name = spec

            # シート名の正規化
            sheet_name = sheet_name.upper()
            # 実際のシート名を検索（大文字小文字対応）
            actual_sheet = None
            for sn in wb.sheetnames:
                if sn.upper() == sheet_name:
                    actual_sheet = sn
                    break

            if not actual_sheet:
                print(f"  警告: シート「{sheet_name}」が見つかりません。スキップ。")
                continue

            fields, rows, add_data, key_data, is_primary_key = read_sheet_data(
                wb, actual_sheet, target_color, target_columns
            )

            if fields is None:
                continue

            if all_fields is None:
                all_fields = fields

            if rows:
                all_rows.extend(rows)

        if all_fields is None:
            print(f"  警告: データが取得できませんでした。空のJSONを出力します。")
            json_str = '{"data":[]}'
            save_file(json_dir, f"{output_file_name}.json", json_str)
            row += 1
            continue

        # JSON生成・保存
        json_str = generate_json(all_fields, all_rows)
        json_path = save_file(json_dir, f"{output_file_name}.json", json_str)
        print(f"  JSON出力: {json_path}")

        # C#生成・保存
        cs_str = generate_csharp(
            all_fields, all_rows,
            actual_sheet, output_file_name,
            key_data, is_primary_key, add_data,
            EXCEL_FILE
        )
        cs_filename = to_pascal_case(output_file_name) + ".cs"
        cs_path = save_file(json_dir, cs_filename, cs_str)
        print(f"  C#出力:   {cs_path}")

        processed += 1
        row += 1

    wb.close()

    print(f"\n===== 完了 =====")
    print(f"処理: {processed}件, スキップ: {skipped}件")
    print(f"出力先: {json_dir}")


if __name__ == "__main__":
    main()


"""
Excel(.xlsm) → C# IDクラス 変換ツール
元のVBAマクロ「IDクラス出力ツール」のPython版

使い方:
  python generate_id.py
"""

import os
import re
import warnings

import openpyxl

warnings.filterwarnings("ignore", category=UserWarning)

# ====== 設定 ======
EXCEL_FILE = "データ.xlsm"
TOOL_SHEET = "IDクラス出力ツール"
START_MARKER = "ここから"
END_MARKER = "ここまで"
COMMENT_PREFIX = "//"
# ==================


# ============================================================
#  色の取得・比較（generate.py と共通）
# ============================================================
def get_cell_color(cell):
    """セルの背景色を比較可能な形式で返す"""
    fill = cell.fill
    if fill is None or fill.patternType is None:
        return None
    fg = fill.fgColor
    if fg is None or fg.type is None:
        return None

    if fg.type == "rgb" and fg.rgb and fg.rgb != "00000000":
        return ("rgb", str(fg.rgb))
    elif fg.type == "theme":
        theme = fg.theme
        tint = round(fg.tint, 4) if fg.tint else 0.0
        return ("theme", theme, tint)
    elif fg.type == "indexed":
        return ("indexed", fg.indexed)
    return None


# ============================================================
#  名前の整形（VBAの ReplaceAndProper + IdGenerate 相当）
# ============================================================
def replace_and_proper(s):
    """
    VBAの ReplaceAndProper 相当。
    区切り文字（スペース）の直後にある2文字以上の英字連続を ProperCase にし、
    スペースを除去して返す。
    
    VBAの処理:
      1. 先頭にスペースを付加
      2. 正規表現 " [a-z]{2,}| [A-Z]{2,}" にマッチした部分を ProperCase に
      3. スペースを除去
    """
    s = " " + s
    # VBAと同じパターン: スペース + 小文字2文字以上 or スペース + 大文字2文字以上
    pattern = r' [a-z]{2,}| [A-Z]{2,}'

    def to_proper(match):
        text = match.group(0)
        # ProperCase: 先頭大文字 + 残り小文字（スペースはそのまま保持）
        return text[0] + text[1].upper() + text[2:].lower()

    s = re.sub(pattern, to_proper, s)
    s = s.replace(" ", "")
    return s


def id_generate(name, id_val):
    """
    VBAの IdGenerate 相当。
    名前を整形し、C#のconst定義文字列を返す。
    """
    # 改行を除去
    name = name.replace("\r", "").replace("\n", "")

    # 区切り文字をスペースに変換
    name = name.replace("_", " ")
    name = name.replace(".", " ")
    name = name.replace("/", " ")
    name = name.replace("\\", " ")

    # ProperCase化 + スペース除去
    name = replace_and_proper(name)

    return f"\t\tpublic const int {name} = {id_val};"


# ============================================================
#  C#クラスファイル生成・保存
# ============================================================
def save_class_file(directory, filename, ids, book_name):
    """VBAの SaveClassFile 相当"""
    os.makedirs(directory, exist_ok=True)
    out_path = os.path.join(directory, f"{filename}.cs")

    lines = []
    lines.append("//------------------------------------------------------------------------------")
    lines.append(f"// Generated by ExcelToPython From {book_name}")
    lines.append("//------------------------------------------------------------------------------")
    lines.append("")
    lines.append("namespace Define")
    lines.append("{")
    lines.append(f"\tpublic class {filename}")
    lines.append("\t{")

    for id_line in ids:
        lines.append(id_line)

    lines.append("\t}")
    lines.append("}")

    with open(out_path, "w", encoding="utf-8-sig") as f:
        f.write("\n".join(lines) + "\n")

    return out_path


# ============================================================
#  列文字 → 列番号 変換
# ============================================================
def column_letter_to_index(letter):
    """A→1, B→2, ... のように列文字を列番号に変換"""
    return openpyxl.utils.column_index_from_string(letter.strip().upper())


# ============================================================
#  メイン処理
# ============================================================
def main():
    print(f"ファイルを読み込み中: {EXCEL_FILE}")
    wb = openpyxl.load_workbook(EXCEL_FILE, data_only=True)

    if TOOL_SHEET not in wb.sheetnames:
        print(f"エラー: シート「{TOOL_SHEET}」が見つかりません。")
        return

    ws = wb[TOOL_SHEET]
    base_dir = os.path.dirname(os.path.abspath(EXCEL_FILE))
    id_dir = os.path.join(base_dir, "id")

    row = 2
    processed = 0
    skipped = 0

    while ws.cell(row=row, column=1).value is not None:
        a_val = str(ws.cell(row=row, column=1).value).strip()
        b_val = str(ws.cell(row=row, column=2).value).strip() if ws.cell(row=row, column=2).value else ""

        # E列に色がついていたらスキップ（VBAの skipSettingColumn = "E"）
        e_color = get_cell_color(ws.cell(row=row, column=5))
        if e_color is not None:
            print(f"  スキップ: {a_val}（E列に色あり）")
            row += 1
            skipped += 1
            continue

        # 設定の取得
        load_settings = a_val.split("+")

        b_clean = b_val.replace("\r", "").replace("\n", "").replace(" ", "")
        filenames = b_clean.split("+")
        output_file_name = filenames[0]

        # C列: 名前列、D列: ID列
        name_col_letter = str(ws.cell(row=row, column=3).value).strip() if ws.cell(row=row, column=3).value else "B"
        id_col_letter = str(ws.cell(row=row, column=4).value).strip() if ws.cell(row=row, column=4).value else "A"

        name_col = column_letter_to_index(name_col_letter)
        id_col = column_letter_to_index(id_col_letter)

        print(f"\n処理中: {a_val} → {output_file_name}")

        all_ids = []

        for setting in load_settings:
            sheet_name = setting.strip().replace("\r", "").replace("\n", "").replace(" ", "").upper()

            # 実際のシート名を検索（大文字小文字対応）
            actual_sheet = None
            for sn in wb.sheetnames:
                if sn.upper() == sheet_name:
                    actual_sheet = sn
                    break

            if not actual_sheet:
                print(f"  警告: シート「{sheet_name}」が見つかりません。スキップ。")
                continue

            data_ws = wb[actual_sheet]

            # 「ここから」「ここまで」を探す
            start_row = None
            end_row = None
            for r in range(1, data_ws.max_row + 1):
                val = data_ws.cell(row=r, column=1).value
                if val == START_MARKER and start_row is None:
                    start_row = r
                elif val == END_MARKER and start_row is not None:
                    end_row = r
                    break

            if start_row is None:
                print(f"  警告: シート「{actual_sheet}」に「{START_MARKER}」が見つかりません。")
                continue
            if end_row is None:
                print(f"  警告: シート「{actual_sheet}」に「{END_MARKER}」が見つかりません。")
                continue

            # VBA: startRow + 2 からデータ開始（ここから + ヘッダ行 + 定義行）
            data_start_row = start_row + 3

            for r in range(data_start_row, end_row + 1):
                name_val = data_ws.cell(row=r, column=name_col).value
                id_val = data_ws.cell(row=r, column=id_col).value

                # 名前が空でない、かつIDが数値の場合のみ出力
                if name_val is not None and str(name_val).strip() != "":
                    id_str = str(id_val) if id_val is not None else ""
                    # IsNumeric チェック相当
                    try:
                        int(float(id_str))
                        is_numeric = True
                    except (ValueError, TypeError):
                        is_numeric = False

                    if is_numeric:
                        line = id_generate(str(name_val), str(int(float(id_str))))
                        all_ids.append(line)

        # ファイル保存
        out_path = save_class_file(id_dir, output_file_name, all_ids, EXCEL_FILE)
        print(f"  C#出力: {out_path}")

        processed += 1
        row += 1

    wb.close()

    print(f"\n===== 完了 =====")
    print(f"処理: {processed}件, スキップ: {skipped}件")
    print(f"出力先: {id_dir}")


if __name__ == "__main__":
    main()
